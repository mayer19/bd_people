import openpyxl
from openpyxl import Workbook
def exel_db():
    #Read the exel
    wb_object = openpyxl.load_workbook("BD_test.xlsx")
    sheet_obj = wb_object.active
    cell_obj = sheet_obj.cell(row=1, column=1)
    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column
    #print("Numero de linhas: {}".format(max_row))
    #print("Numero de colunas: {}".format(max_col))

    #Loop to read the columns and send to lists
    name_list = []
    lastname_list = []
    fullname_list = []
    email_list = []
    emailpessoal_list = []
    job_list = []
    company_list = []
    phone_list = []
    fixphone_list = []
    linkedin_list = []
    id_list = []
    sector_list = []
    subsector_list = []
    zone_list = []
    notes_list = []
    a_list = []

    rows = sheet_obj.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)

    for name,last_name,full_name,email,email_pessoal,job,company,phone,fix_phone,linkedin,id,sector,subsector,zone,notes,a in rows:
        #if name.value or last_name.value or full_name.value or email.value or email_pessoal.value or job.value or company.value or phone.value or fix_phone.value or linkedin.value or id.value or sector.value or subsector.value or zone.value or notes.value or a.value == None:
        if name.value == None:
            name_list.append('')
        else:
            name_list.append(name.value)
        if last_name.value == None:
            lastname_list.append('')
        else:
            lastname_list.append(last_name.value)
        if full_name == None:
            fullname_list.append('')
        else:
            fullname_list.append(full_name.value)
        if email.value == None:
            email_list.append('')
        else:
            email_list.append(email.value)
        if email_pessoal == None:
            emailpessoal_list.append('')
        else:
            emailpessoal_list.append(email_pessoal.value)
        if job.value == None:
            job_list.append('')
        else:
            job_list.append(job.value)
        if company.value == None:
            company_list.append('')
        else:
            company_list.append(company.value)
        if phone.value == None:
            phone_list.append('')
        else:
            phone_list.append(phone.value)
        if fix_phone == None:
            fixphone_list.append('')
        else:
            fixphone_list.append(fix_phone)
        if linkedin.value == None:
            linkedin_list.append('')
        else:
            linkedin_list.append(linkedin.value)
        if id.value == None:
            id_list.append('')
        else:
            id_list.append(id.value)
        if sector.value == None:
            sector_list.append('')
        else:
            sector_list.append(sector.value)
        if subsector.value == None:
            subsector_list.append('')
        else:
            subsector_list.append(subsector.value)
        if zone.value == None:
            zone_list.append('')
        else:
            zone_list.append(zone.value)
        if notes.value == None:
            notes_list.append('')
        else:
            notes_list.append(notes.value)

    print('COPIADO DO EXCEL COM SUCESSO.')

exel_db()